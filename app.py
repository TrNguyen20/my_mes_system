import io
import csv
import openpyxl
import gspread
from datetime import datetime
from functools import wraps
from flask import Flask, render_template, request, redirect, Response, flash, url_for, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from flask_bcrypt import Bcrypt
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy import or_

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///mes_v3.db'
app.config['SECRET_KEY'] = 'enterprise_mes_master_2025'

db = SQLAlchemy(app)
bcrypt = Bcrypt(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'

# --- DATABASE MODELS ---

class User(db.Model, UserMixin):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(128), nullable=False)
    role = db.Column(db.String(20), default='operator') # admin, manager, operator
    is_active = db.Column(db.Boolean, default=True)
    date_created = db.Column(db.DateTime, default=datetime.now)

    def set_password(self, password):
        self.password_hash = bcrypt.generate_password_hash(password).decode('utf-8')

    def check_password(self, password):
        return bcrypt.check_password_hash(self.password_hash, password)

class WorkOrder(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    po_number = db.Column(db.String(50), unique=True, nullable=False)
    part_name = db.Column(db.String(100), nullable=False)
    quantity = db.Column(db.Integer, nullable=False)
    status = db.Column(db.String(20), default='Pending') # Pending, In Progress, Completed
    current_step_index = db.Column(db.Integer, default=0)
    date_created = db.Column(db.DateTime, default=datetime.now)
    date_started = db.Column(db.DateTime, nullable=True)
    date_completed = db.Column(db.DateTime, nullable=True)
    operations = db.relationship('Operation', backref='work_order', lazy=True, cascade="all, delete-orphan")

    def get_duration(self):
        if self.date_started and self.date_completed:
            diff = self.date_completed - self.date_started
            return round(diff.total_seconds() / 60, 2)
        return None

class Operation(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    work_order_id = db.Column(db.Integer, db.ForeignKey('work_order.id'), nullable=False)
    name = db.Column(db.String(100), nullable=False)
    sequence = db.Column(db.Integer, nullable=False)
    status = db.Column(db.String(20), default='Pending') # Pending, In Progress, Completed
    start_time = db.Column(db.DateTime, nullable=True)
    end_time = db.Column(db.DateTime, nullable=True)

    def get_op_duration(self):
        if self.start_time and self.end_time:
            diff = self.end_time - self.start_time
            return round(diff.total_seconds() / 60, 2)
        return 0

class AuditLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    action = db.Column(db.String(50), nullable=False)
    target_type = db.Column(db.String(50), nullable=False)
    target_id = db.Column(db.Integer, nullable=False)
    detail = db.Column(db.String(200))
    timestamp = db.Column(db.DateTime, default=datetime.now)
    user = db.relationship('User', backref='logs')

# --- HELPERS & TWO-WAY SYNC ---

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated or current_user.role not in roles:
                flash("Quyền truy cập bị từ chối!", "danger")
                return redirect(url_for('index'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

def log_action(action, target_type, target_id, detail=''):
    try:
        log = AuditLog(user_id=current_user.id if current_user.is_authenticated else 1, 
                       action=action, target_type=target_type, target_id=target_id, detail=detail)
        db.session.add(log)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f"Audit log error: {e}")

def sync_to_google_sheets():
    """Đồng bộ 2 chiều: Đẩy toàn bộ trạng thái DB hiện tại lên Google Sheets"""
    try:
        SHEET_ID = "1VeoJY4tW3EoN-IK_kXKGlExH_E1Jzr6ffQB-eVbJO1w"
        gc = gspread.service_account(filename='credentials.json')
        sh = gc.open_by_key(SHEET_ID)
        worksheet = sh.get_worksheet(0)
        
        data_to_sync = [["po_number", "part_name", "quantity", "status", "current_step"]]
        orders = WorkOrder.query.all()
        for o in orders:
            ops = sorted(o.operations, key=lambda x: x.sequence)
            current_op = ops[o.current_step_index-1].name if (o.current_step_index > 0 and len(ops) >= o.current_step_index) else "N/A"
            data_to_sync.append([o.po_number, o.part_name, o.quantity, o.status, current_op])
        
        worksheet.clear()
        worksheet.update('A1', data_to_sync)
    except Exception as e:
        print(f"Two-Way Sync Error: {e}")

def fetch_sheet_orders(sheet_url_or_id):
    try:
        SHEET_ID = "1VeoJY4tW3EoN-IK_kXKGlExH_E1Jzr6ffQB-eVbJO1w"
        gc = gspread.service_account(filename='credentials.json')
        sh = gc.open_by_key(SHEET_ID)
        worksheet = sh.get_worksheet(0)
        return worksheet.get_all_records()
    except Exception as e:
        print(f"Google Sheets Fetch Error: {e}")
        return []

# --- MAIN ROUTES ---

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        user = User.query.filter_by(username=request.form['username']).first()
        if user and user.check_password(request.form['password']):
            if not user.is_active:
                flash("Tài khoản đã bị vô hiệu hóa!", "danger")
                return redirect(url_for('login'))
            login_user(user, remember=request.form.get('remember'))
            return redirect(url_for('index'))
        flash('Sai tài khoản hoặc mật khẩu!', 'danger')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/')
@login_required
def index():
    search = request.args.get('search', '')
    query = WorkOrder.query
    if search:
        query = query.filter(or_(
            WorkOrder.po_number.contains(search),
            WorkOrder.part_name.contains(search),
            WorkOrder.status.contains(search),
            db.cast(WorkOrder.quantity, db.String).contains(search)
        ))
    orders = query.order_by(WorkOrder.date_created.desc()).all()
    stats = {
        'pending': WorkOrder.query.filter_by(status='Pending').count(),
        'progress': WorkOrder.query.filter_by(status='In Progress').count(),
        'completed': WorkOrder.query.filter_by(status='Completed').count(),
    }
    return render_template('index.html', orders=orders, stats=stats, search_val=search)

@app.route('/add', methods=['POST'])
@login_required
@role_required('admin', 'manager')
def add_order():
    po = request.form['po_number'].strip()
    if WorkOrder.query.filter_by(po_number=po).first():
        flash(f"Lỗi: Mã PO {po} đã tồn tại!", "danger")
        return redirect(url_for('index'))
    
    new_order = WorkOrder(po_number=po, part_name=request.form['part_name'], quantity=request.form['quantity'])
    db.session.add(new_order)
    db.session.flush()

    ops_list = request.form.getlist('operations[]')
    if not ops_list or ops_list[0] == "": ops_list = ["Sản xuất chung"]

    for i, op_name in enumerate(ops_list):
        db.session.add(Operation(work_order_id=new_order.id, name=op_name, sequence=i+1))
    
    db.session.commit()
    log_action('CREATE_ROUTING', 'WorkOrder', new_order.id, f'PO: {po}')
    sync_to_google_sheets()
    flash("Đã mở lệnh sản xuất kèm quy trình công nghệ!", "success")
    return redirect(url_for('index'))

@app.route('/edit/<int:id>', methods=['POST'])
@login_required
@role_required('admin', 'manager')
def edit_order(id):
    order = WorkOrder.query.get_or_404(id)
    old_po = order.po_number
    order.po_number = request.form['po_number'].strip()
    order.part_name = request.form['part_name'].strip()
    order.quantity = int(request.form['quantity'])
    order.status = request.form['status']
    order.current_step_index = int(request.form['current_step_index'])
    
    db.session.commit()
    log_action('EDIT_FULL', 'WorkOrder', id, f'Edited PO {old_po} -> {order.po_number} (Step {order.current_step_index})')
    sync_to_google_sheets()
    flash("Cập nhật toàn diện thông tin & quy trình thành công!", "success")
    return redirect(url_for('index'))

@app.route('/update/<int:id>')
@login_required
def update_status(id):
    order = WorkOrder.query.get_or_404(id)
    ops = sorted(order.operations, key=lambda x: x.sequence)
    
    if order.status == 'Pending':
        order.status = 'In Progress'
        order.date_started = datetime.now()
        order.current_step_index = 1
        if ops:
            ops[0].status = 'In Progress'
            ops[0].start_time = datetime.now()
    elif order.status == 'In Progress':
        current_op = ops[order.current_step_index - 1]
        current_op.status = 'Completed'
        current_op.end_time = datetime.now()
        
        if order.current_step_index < len(ops):
            order.current_step_index += 1
            next_op = ops[order.current_step_index - 1]
            next_op.status = 'In Progress'
            next_op.start_time = datetime.now()
        else:
            order.status = 'Completed'
            order.date_completed = datetime.now()
            
    db.session.commit()
    log_action('STEP_ADVANCE', 'WorkOrder', order.id, f'Advanced to step {order.current_step_index}')
    sync_to_google_sheets()
    return redirect(url_for('index'))

@app.route('/delete/<int:id>')
@login_required
@role_required('admin')
def delete_order(id):
    order = WorkOrder.query.get_or_404(id)
    po_ref = order.po_number
    db.session.delete(order)
    db.session.commit()
    log_action('DELETE', 'WorkOrder', id, f'Deleted PO: {po_ref}')
    sync_to_google_sheets()
    flash(f'Đã xóa lệnh {po_ref}', 'warning')
    return redirect(url_for('index'))

# --- SYNC & EXPORT ---

@app.route('/import-from-sheets')
@login_required
@role_required('admin', 'manager')
def import_from_sheets():
    sheet_data = fetch_sheet_orders("1VeoJY4tW3EoN-IK_kXKGlExH_E1Jzr6ffQB-eVbJO1w")
    imported = 0
    for row in sheet_data:
        po = str(row.get('po_number', '')).strip()
        if po and not WorkOrder.query.filter_by(po_number=po).first():
            new_o = WorkOrder(po_number=po, part_name=str(row.get('part_name', 'N/A')), quantity=int(row.get('quantity', 0)))
            db.session.add(new_o)
            db.session.flush()
            db.session.add(Operation(work_order_id=new_o.id, name="Gia công chuẩn", sequence=1))
            imported += 1
    db.session.commit()
    log_action('SYNC_SHEETS_PULL', 'System', 0, f'Pulled {imported} orders')
    sync_to_google_sheets()
    flash(f"📊 Đồng bộ Google Sheets thành công: +{imported} đơn mới", "success")
    return redirect(url_for('index'))

@app.route('/download-excel-template')
def download_template():
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Import Template"
    
    header_style = PatternFill(start_color="2d3436", end_color="2d3436", fill_type="solid")
    sample_style = PatternFill(start_color="dfe6e9", end_color="dfe6e9", fill_type="solid")
    white_font = Font(bold=True, color="FFFFFF")
    warning_font = Font(italic=True, color="e17055")
    center_align = Alignment(horizontal="center", vertical="center")

    headers = ["po_number", "part_name", "quantity"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_style; cell.font = white_font; cell.alignment = center_align

    for r in range(20, 26):
        ws.cell(row=r, column=1, value=f"PO-SAMPLE-{r}").fill = sample_style
        ws.cell(row=r, column=2, value="Part Name Example").fill = sample_style
        ws.cell(row=r, column=3, value=150).fill = sample_style

    ws.merge_cells('A6:C6')
    ws['A6'] = "⚠ Delete sample rows before importing. Keep header row."
    ws['A6'].font = warning_font; ws['A6'].alignment = center_align
    ws.freeze_panes = "A2"
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name="MES_Template.xlsx")

@app.route('/import-excel', methods=['POST'])
@login_required
@role_required('admin', 'manager')
def import_excel():
    file = request.files.get('excel_file')
    if not file: return redirect(url_for('index'))
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    imported = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        po = str(row[0]).strip() if row[0] else ""
        if not po or po == "None" or WorkOrder.query.filter_by(po_number=po).first(): continue
        new_o = WorkOrder(po_number=po, part_name=str(row[1]), quantity=int(row[2]))
        db.session.add(new_o); db.session.flush()
        db.session.add(Operation(work_order_id=new_o.id, name="Gia công cơ khí", sequence=1))
        imported += 1
    db.session.commit()
    log_action('IMPORT_EXCEL', 'System', 0, f'Imported {imported}')
    sync_to_google_sheets()
    flash(f"✅ Đã nhập {imported} đơn hàng từ Excel", "success")
    return redirect(url_for('index'))

@app.route('/export-csv')
@login_required
@role_required('admin', 'manager')
def export_csv():
    output = io.StringIO()
    output.write('\ufeff')
    writer = csv.writer(output)
    writer.writerow(['PO Number', 'Tên sản phẩm', 'Số lượng', 'Trạng thái', 'Lead Time (Min)'])
    for o in WorkOrder.query.all():
        writer.writerow([o.po_number, o.part_name, o.quantity, o.status, o.get_duration()])
    output.seek(0)
    return Response(output.getvalue(), mimetype="text/csv", headers={"Content-Disposition":"attachment;filename=report_mes.csv"})

# --- ADMIN USER MANAGEMENT & AUDIT ---

@app.route('/admin/users')
@login_required
@role_required('admin')
def admin_users():
    users = User.query.all()
    user_count = User.query.count()
    return render_template('admin_users.html', users=users, user_count=user_count)

@app.route('/admin/create-user', methods=['POST'])
@login_required
@role_required('admin')
def admin_create_user():
    user = User(username=request.form['username'], email=request.form['email'], role=request.form['role'])
    user.set_password(request.form['password'])
    db.session.add(user); db.session.commit()
    log_action('CREATE_USER', 'User', user.id, f'Created {user.username}')
    flash("Tạo người dùng thành công!", "success")
    return redirect(url_for('admin_users'))

@app.route('/admin/toggle-user/<int:id>', methods=['POST'])
@login_required
@role_required('admin')
def admin_toggle_user(id):
    if current_user.id != id:
        u = User.query.get_or_404(id)
        u.is_active = not u.is_active
        db.session.commit()
        log_action('TOGGLE_USER', 'User', u.id, f'Toggle active to {u.is_active}')
    return redirect(url_for('admin_users'))

@app.route('/audit-log')
@login_required
@role_required('admin')
def view_audit_log():
    logs = AuditLog.query.order_by(AuditLog.timestamp.desc()).limit(100).all()
    return render_template('audit_log.html', logs=logs)

# --- CLI COMMANDS ---

@app.cli.command("seed-admin")
def seed_admin():
    db.create_all()
    if not User.query.filter_by(username='admin').first():
        u = User(username='admin', email='admin@mes.com', role='admin')
        u.set_password('admin123')
        db.session.add(u); db.session.commit()
        print("Admin account created: admin / admin123")

if __name__ == "__main__":
    with app.app_context():
        db.create_all()
    app.run(debug=True)